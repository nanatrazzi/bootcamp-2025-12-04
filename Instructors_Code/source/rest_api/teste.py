# app.py
import pandas as pd
from openpyxl import load_workbook

file_path = "Simulador-2025-Calculo-de-Contribuicao-UR-1(1).xlsx"
# pd.set_option("display.max_colwidth", None)

wb_formula = load_workbook(file_path, data_only=False)
wb_values = load_workbook(file_path, data_only=True)

sheet_f = wb_formula.active
sheet_v = wb_values.active

linhas = []
for row_f, row_v in zip(sheet_f.iter_rows(), sheet_v.iter_rows()):
    for cf, cv in zip(row_f, row_v):

        formula = cf.value if isinstance(cf.value, str) and cf.value.startswith("=") else None
        valor = cv.value

        linhas.append({
            "celula": cf.coordinate,
            "valor_calculado": valor,
            "formula": formula
        })

df = pd.DataFrame(linhas)

df = df[~(df["valor_calculado"].isna() & df["formula"].isna())]

print(df)

# print(df["formula"].dropna())
# for index, row in df[df["formula"].notna()].iterrows():
#     print(f"{row['celula']} -> {row['formula']}")



